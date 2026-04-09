from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook

# -----------------------------
# CONFIG
# -----------------------------
RAW_DIR = Path("data/raw") 
CALLIGRAPHY_TYPES = ["encadenada", "italica_cursiva", "procesal", "redonda"]


# -----------------------------
# Utilities
# -----------------------------

def index_txt_files(directory: Path):
    return list(directory.rglob("*.txt"))


def build_htr_prefix_index(htr_files_by_style):
    index = defaultdict(list)

    for style, files in htr_files_by_style.items():
        for htr in files:
            stem = htr.name[:-8]  # remove _HTR.txt
            parts = stem.split("_")

            for i in range(1, len(parts) + 1):
                prefix = "_".join(parts[:i])
                index[prefix].append({
                    "style": style,
                    "path": htr
                })

    return index


# -----------------------------
# Match report
# -----------------------------

htr_files = {
    s: index_txt_files(RAW_DIR / s)
    for s in CALLIGRAPHY_TYPES
}

gt_files = index_txt_files(RAW_DIR / "ground_truths")

total_htr = sum(len(v) for v in htr_files.values())
total_gt = len(gt_files)

htr_index = build_htr_prefix_index(htr_files)

multiple = {}
zero = []
single = 0

for gt in gt_files:
    gt_stem = gt.name[:-7]  # remove _GT.txt
    matches = htr_index.get(gt_stem, [])

    if len(matches) == 0:
        zero.append(gt.name)
    elif len(matches) == 1:
        single += 1
    else:
        multiple[gt.name] = [
            (m["path"].name, m["style"])
            for m in matches
        ]

# -----------------------------
# Console Summary
# -----------------------------

print("\n=== MATCH SUMMARY ===")
print(f"Total GT files: {total_gt}")
print(f"Total HTR files: {total_htr}")
print(f"Single matches: {single}")
print(f"Zero matches: {len(zero)}")
print(f"Multiple matches: {len(multiple)}")


# -----------------------------
# Create Excel workbook
# -----------------------------

wb = Workbook()

# -----------------------------
# Summary Sheet
# -----------------------------
ws_summary = wb.active
ws_summary.title = "Summary"

ws_summary.append(["Metric", "Count"])
ws_summary.append(["Total GT files", total_gt])
ws_summary.append(["Total HTR files", total_htr])
ws_summary.append(["Single matches", single])
ws_summary.append(["Zero matches", len(zero)])
ws_summary.append(["Multiple matches", len(multiple)])

# -----------------------------
# Multiple Matches Sheet
# -----------------------------
ws_multi = wb.create_sheet(title="Multiple_Matches")
ws_multi.append(["GT Filename", "HTR Filename", "Style"])

for gt_name in sorted(multiple.keys()):
    for htr_name, style in multiple[gt_name]:
        ws_multi.append([gt_name, htr_name, style])

# -----------------------------
# Zero Matches Sheet
# -----------------------------
ws_zero = wb.create_sheet(title="Zero_Matches")
ws_zero.append(["GT Filename"])

for filename in sorted(zero):
    ws_zero.append([filename])

# -----------------------------
# Save Excel (in project root)
# -----------------------------
output_path = Path("htr_gt_match_report.xlsx")
wb.save(output_path)

print(f"\nWritten to: {output_path.resolve()}")